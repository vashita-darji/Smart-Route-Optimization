import requests
import sys
def get_traffic_info(lat1, lon1, lat2, lon2):
    url = f'https://api.mapbox.com/directions/v5/mapbox/driving/{lon1},{lat1};{lon2},{lat2}?alternatives=false&geometries=geojson&steps=false&access_token=pk.eyJ1IjoiZGVla3NoaXRhMTYwOCIsImEiOiJjbTZrc21iNjAwMWV3MmpxeXg4cmZ4djJnIn0.BcVfHhiyldVFKnkqbCqm3Q'
    response = requests.get(url)
    data = response.json()

    # Example: Extract travel time (seconds)
    travel_time = data['routes'][0]['duration']

    # Classify traffic
    if travel_time < 1800:
        return 2
    elif travel_time < 3600:
        return 1
    else:
        return 0

import pandas as pd
import random
from geopy.distance import geodesic

# Haversine function to calculate distance
def haversine(lat1, lon1, lat2, lon2):
    point1 = (lat1, lon1)
    point2 = (lat2, lon2)
    return geodesic(point1, point2).km

# Coordinates of the store
store_lat = 19.075887
store_lon = 72.877911

# Read your CSV data (replace 'your_file.csv' with the path to your file)
df = pd.read_excel("Shipment_Data.xlsx")

# Calculate distance and append it to the DataFrame
df['Distance from Store (km)'] = df.apply(lambda row: haversine(store_lat, store_lon, row['Latitude'], row['Longitude']), axis=1)

# Sort the data by the timeslot (this will sort the rows by the 'Delivery Timeslot' column)
df_sorted = df.sort_values(by='Delivery Timeslot')

# Separate the data by timeslot
timeslot_data = {}
for timeslot in df_sorted['Delivery Timeslot'].unique():
    timeslot_data[timeslot] = df_sorted[df_sorted['Delivery Timeslot'] == timeslot]

# Define the vehicles with their types, number, capacity, and trip radius
vehicles = [
    {'Vehicle Type': '3W', 'Number': 50, 'Shipments_Capacity': 5, 'Max Trip Radius': 15},
    {'Vehicle Type': '4W-EV', 'Number': 25, 'Shipments_Capacity': 8, 'Max Trip Radius': 20},
    {'Vehicle Type': '4W', 'Number': 'Any', 'Shipments_Capacity': 25, 'Max Trip Radius': 'Any'}
]

# Function to calculate and assign vehicles to shipments based on priority and vehicle properties
def assign_vehicles_to_shipments(timeslot_df, used_vehicles_count):
    assigned_vehicles = []
    remaining_shipments = timeslot_df.copy()
    available_vehicles = []

    # Add priority vehicles (3W and 4W-EV) first
    for vehicle in vehicles[:2]:  # First two vehicles: '3W' and '4W-EV'
        for _ in range(vehicle['Number']):
            available_vehicles.append({'Vehicle Type': vehicle['Vehicle Type'],
                                       'Capacity': vehicle['Shipments_Capacity'],
                                       'Max Trip Radius': vehicle['Max Trip Radius']})

    # Add '4W' vehicles only if needed (or if unlimited availability)
    for vehicle in vehicles[2:]:  # '4W' vehicle
        if vehicle['Number'] != 'Any':  # Limit to available vehicles if not 'Any'
            for _ in range(vehicle['Number']):
                available_vehicles.append({'Vehicle Type': vehicle['Vehicle Type'],
                                           'Capacity': vehicle['Shipments_Capacity'],
                                           'Max Trip Radius': vehicle['Max Trip Radius']})
        else:
            # '4W' has infinite availability, add based on need
            available_vehicles.append({'Vehicle Type': vehicle['Vehicle Type'],
                                       'Capacity': vehicle['Shipments_Capacity'],
                                       'Max Trip Radius': vehicle['Max Trip Radius']})

    unassigned_shipments = []  # Track unassigned shipments
    random.shuffle(available_vehicles)

    # Assign vehicles to shipments respecting the vehicle's capacity and max trip radius
    for vehicle in available_vehicles:
        if not remaining_shipments.empty:
            current_capacity = 0  # Number of shipments assigned to this vehicle
            assigned_shipments = []
            trip_distance = 0  # Initialize total trip distance
            prev_lat, prev_long = store_lat, store_lon
            remaining_shipments_list = remaining_shipments.copy()

            for _, shipment in remaining_shipments_list.iterrows():
                # Calculate the distance for the current shipment
                shipment_distance = haversine(prev_lat, prev_long, shipment['Latitude'], shipment['Longitude'])
                return_distance = haversine(shipment['Latitude'], shipment['Longitude'], store_lat, store_lon)
                total_trip_distance = trip_distance + shipment_distance + return_distance

                # Check if the shipment can be assigned to the current vehicle
                if current_capacity + 1 <= vehicle['Capacity']:
                    if vehicle['Max Trip Radius'] != 'Any':  # Check for max trip radius only if it's not 'Any'
                        if total_trip_distance <= vehicle['Max Trip Radius']:
                            current_capacity += 1
                            assigned_shipments.append(shipment)
                            trip_distance += shipment_distance
                            remaining_shipments = remaining_shipments.drop(shipment.name)
                            prev_lat, prev_long = shipment['Latitude'], shipment['Longitude']
                    else:
                        # No max radius constraint, just assign the shipment
                        current_capacity += 1
                        assigned_shipments.append(shipment)
                        trip_distance += shipment_distance
                        remaining_shipments = remaining_shipments.drop(shipment.name)
                        prev_lat, prev_long = shipment['Latitude'], shipment['Longitude']

            if assigned_shipments:
                assigned_vehicles.append((
                    vehicle['Vehicle Type'],
                    [s['Shipment ID'] for s in assigned_shipments],
                    trip_distance,
                    current_capacity / vehicle['Capacity'],
                    get_traffic_info(store_lat,store_lon,assigned_shipments[-1]['Latitude'], assigned_shipments[-1]['Longitude'])
                ))

                # Update the used vehicle count
                used_vehicles_count[vehicle['Vehicle Type']] += 1
            else:
                # If no shipments were assigned to the vehicle, track it as unassigned
                unassigned_shipments.extend(remaining_shipments['Shipment ID'].tolist())

    # If any shipments are still unassigned, assign them to '4W' vehicles, but limit to 25 shipments per vehicle
    if not remaining_shipments.empty:
        # Group remaining shipments into batches of 25 for the 4W vehicles
        remaining_shipments_list = remaining_shipments.copy()
        while not remaining_shipments_list.empty:
            current_capacity = 0
            assigned_shipments = []
            for _, shipment in remaining_shipments_list.iterrows():
                if current_capacity < 25:  # '4W' can hold up to 25 shipments
                    assigned_shipments.append(shipment)
                    current_capacity += 1
                    remaining_shipments_list = remaining_shipments_list.drop(shipment.name)
                else:
                    break

            # Now assign this batch to a new 4W vehicle
            assigned_vehicles.append(('4W', [s['Shipment ID'] for s in assigned_shipments], 0, current_capacity / 25, get_traffic_info(store_lat,store_lon,assigned_shipments[-1]['Latitude'], assigned_shipments[-1]['Longitude'])))

    # Return assigned vehicles and unassigned shipments
    return assigned_vehicles, used_vehicles_count, remaining_shipments_list

# Function to generate initial population
def generate_init_population(num_chromosomes, slot):

    per_timeslot = num_chromosomes  # Use the full number of chromosomes for this single timeslot
    population = []
    data = timeslot_data[slot]  # Get data for the selected timeslot

    used_vehicles_per_timeslot = {'3W': 0, '4W-EV': 0, '4W': 0}

    for i in range(per_timeslot):
        print('Generating Chromosome: ',i)
        assigned_vehicles, used_vehicles_per_timeslot, unassigned_shipments = assign_vehicles_to_shipments(
            data, used_vehicles_per_timeslot
        )
        population.append(assigned_vehicles)

    return population

if len(sys.argv) > 1:
    time_slot = sys.argv[1]  # Get the time slot string (e.g., '09:30:00-12:00:00')
else:
    time_slot = '09:30:00-12:00:00'  # Default value if no argument is provided
    print(time_slot)

# Generate the initial population with the selected time slot
population = generate_init_population(100, time_slot)
# Generate the initial population
# population = generate_init_population(2,'09:30:00-12:00:00')

# Print the number of shipments assigned to each chromosome (vehicle set)
for idx, chromosome in enumerate(population):
    total_shipments_assigned = sum(len(vehicle[1]) for vehicle in chromosome)
    print(f"Chromosome {idx + 1} has {total_shipments_assigned} shipments assigned.")

for trip in population[0]:
    print(trip)
#Lower is better
def fitness_function(chromosome):
    total_trip_length = 0
    total_num_trips = 0
    total_capacity_utilization = 0
    penalty = 0

    # Calculate trip lengths, number of trips, and capacity utilization
    for vehicle_assignment in chromosome:
        trip_distance = vehicle_assignment[2]
        num_shipments = len(vehicle_assignment[1])
        capacity_utilization = vehicle_assignment[3]
        traffic=vehicle_assignment[4]
        total_trip_length += trip_distance
        total_num_trips += 1
        total_capacity_utilization += capacity_utilization

        # Penalize trips with less than 50% utilization
        if capacity_utilization < 0.5:
            penalty += 1  # Penalize each trip with less than 50% utilization

    # Calculate average values
    avg_trip_length = total_trip_length / total_num_trips if total_num_trips > 0 else 0
    avg_capacity_utilization = total_capacity_utilization / total_num_trips if total_num_trips > 0 else 0

    return [avg_trip_length, avg_capacity_utilization, total_num_trips, traffic, penalty]
def dominates(individual1, individual2):
    """
    Check if individual1 dominates individual2.
    Assumes the individuals are tuples or lists of fitness values, where lower values are better for the first two
    objectives (minimization) and higher values are better for the last objective (maximization).
    """
    equal_in_all=False
    dominates_atleast_one=False
    fit1 = fitness_function(individual1)
    fit2 = fitness_function(individual2)  # Corrected this line
    if fit1[0] <= fit2[0] and fit1[1] >= fit2[1] and fit1[2] <= fit2[2] and fit1[3]>=fit2[3]:
        equal_in_all = True
    if fit1[0] < fit2[0] or fit1[1] > fit2[1] or fit1[2] < fit2[2] and fit1[3]>fit2[3]:
        dominates_atleast_one = True
    return equal_in_all and dominates_atleast_one


def get_pareto_front(population):
    """
    Find the Pareto front from a population of individuals.
    Each individual is assumed to be a tuple/list of fitness values.
    """
    pareto_front = []

    for i, ind1 in enumerate(population):
        is_dominated = False
        for j, ind2 in enumerate(population):
            if i != j:
                if dominates(ind2, ind1):  # If ind2 dominates ind1
                    is_dominated = True
                    break
        if not is_dominated:
            pareto_front.append(ind1)

    return pareto_front
selected_parents=get_pareto_front(population)
print(len(selected_parents))
import heapq
from geopy.distance import geodesic

def calculate_haversine_distance(shipments):
    """
    Compute the total Haversine distance for a sequence of shipments.
    Each shipment is assumed to have (latitude, longitude) coordinates.
    """
    shipments.insert(0, (store_lat, store_lon))
    shipments.append((store_lat, store_lon))
    total_distance = sum(geodesic(shipments[i], shipments[i + 1]).km for i in range(len(shipments) - 1))
    return total_distance

def weighted_score(trip, distance_weight=0.7, utilization_weight=0.3):
    """
    Compute a weighted score for a trip based on distance and utilization.
    Higher score means a better trip.
    """
    return (1-distance_weight * trip[2]) + (utilization_weight * trip[3]) + trip[4]

def crossover_with_trip_selection(parent1, parent2, max_vehicles):
    """
    Performs crossover by selecting the best trips from parent1 and merging efficient trips from parent2.

    Arguments:
    - parent1, parent2: Lists of trips [(vehicle_type, shipments, trip_distance, utilization)]
    - max_vehicles: Dictionary with max available vehicles per type {"3W": 5, "4W-EV": 8, "4W": 10}

    Returns:
    - offspring: New individual (list of trips)
    """

    # Step 1: Select Top 5 Trips from Parent1
    top_trips = heapq.nlargest(20, parent1, key=weighted_score)
    assigned_shipments = set(shipment for trip in top_trips for shipment in trip[1])

    # Step 2: Process Parent2, Filter Out Already Assigned Shipments
    filtered_trips = []
    max_capacity = {'3W': 5, '4W-EV': 8, '4W': float('inf')}
    max_radius = {'3W': 15, '4W-EV': 20, '4W': float('inf')}

    for trip in parent2:
        vehicle_type, shipments, trip_distance, utilization, traffic = trip
        new_shipments = [s for s in shipments if s not in assigned_shipments]

        if new_shipments:
            new_coords = [(df.loc[df['Shipment ID'] == s, 'Latitude'].values[0],
                           df.loc[df['Shipment ID'] == s, 'Longitude'].values[0]) for s in new_shipments]
            new_trip_distance = calculate_haversine_distance(new_coords)
            new_utilization = min(len(new_shipments) / max_capacity[vehicle_type], 1.0)
            last_shipment_id = new_shipments[-1]
            last_shipment_lat = df.loc[df['Shipment ID'] == last_shipment_id, 'Latitude'].values[0]
            last_shipment_lon = df.loc[df['Shipment ID'] == last_shipment_id, 'Longitude'].values[0]
            new_traffic = get_traffic_info(store_lat, store_lon, last_shipment_lat, last_shipment_lon)
            if new_trip_distance <= max_radius[vehicle_type]:
                filtered_trips.append((vehicle_type, new_shipments, new_trip_distance, new_utilization, new_traffic))
                assigned_shipments.update(new_shipments)

    # Step 3: Greedy Merging of Trips
    merged_trips = []
    considered_trips = set()

    for i, (vehicle_type, shipments, trip_distance, utilization, traffic) in enumerate(filtered_trips):
        if i in considered_trips:
            continue

        best_merge = None
        for j, (vehicle_type2, shipments2, trip_distance2, utilization2, traffic2) in enumerate(filtered_trips):
            if j in considered_trips or i == j or vehicle_type != vehicle_type2:
                continue

            merged_shipments = list(set(shipments + shipments2))
            new_coords = [(df.loc[df['Shipment ID'] == s, 'Latitude'].values[0],
                           df.loc[df['Shipment ID'] == s, 'Longitude'].values[0]) for s in merged_shipments]
            new_trip_distance = calculate_haversine_distance(new_coords)

            if len(merged_shipments) <= max_capacity[vehicle_type] and new_trip_distance <= max_radius[vehicle_type]:
                merged_utilization = len(merged_shipments) / max_capacity[vehicle_type]
                last_shipment_id = merged_shipments[-1]
                last_shipment_lat = df.loc[df['Shipment ID'] == last_shipment_id, 'Latitude'].values[0]
                last_shipment_lon = df.loc[df['Shipment ID'] == last_shipment_id, 'Longitude'].values[0]
                merged_traffic = get_traffic_info(store_lat, store_lon, last_shipment_lat, last_shipment_lon)
                best_merge = (vehicle_type, merged_shipments, new_trip_distance, merged_utilization, merged_traffic)
                considered_trips.update({i, j})
                break

        if best_merge:
            merged_trips.append(best_merge)
        else:
            merged_trips.append((vehicle_type, shipments, trip_distance, utilization, traffic))
            considered_trips.add(i)

    # Step 4: Assign Vehicles and Enforce Constraints
    vehicle_counts = {'3W': 0, '4W-EV': 0, '4W': 0}
    final_trips = []

    for trip in top_trips + merged_trips:
        vehicle_type, shipments, trip_distance, utilization, traffic= trip

        if vehicle_counts[vehicle_type] < max_vehicles[vehicle_type]:
            final_trips.append(trip)
            vehicle_counts[vehicle_type] += 1
        else:
            final_trips.append(('4W', shipments, trip_distance, utilization, traffic))
            vehicle_counts['4W'] += 1

    return final_trips

import random

def mutate_offspring(offspring, mutation_rate):
    """
    Mutates the shipment allocation within trips by shuffling.
    Ensures that mutated trips still meet distance constraints.
    """
    max_radius = {"3W": 15, "4W-EV": 20, "4W": float('inf')}
    new_offspring = []
    mutations_applied = 0
    mutations_failed = 0

    for vehicle_type, shipments, trip_distance, utilization, traffic in offspring:
        if len(shipments) > 1 and random.random() < mutation_rate:  # Apply mutation only if possible
            original_shipments = shipments[:]  # Keep original order for comparison
            random.shuffle(shipments)  # Shuffle shipment order

            # Recalculate trip distance
            new_coords = [(df.loc[df['Shipment ID'] == s, 'Latitude'].values[0],
                           df.loc[df['Shipment ID'] == s, 'Longitude'].values[0])
                          for s in shipments]
            new_trip_distance = calculate_haversine_distance(new_coords)
            last_shipment_id = shipments[-1]
            last_shipment_lat = df.loc[df['Shipment ID'] == last_shipment_id, 'Latitude'].values[0]
            last_shipment_lon = df.loc[df['Shipment ID'] == last_shipment_id, 'Longitude'].values[0]
            new_traffic = get_traffic_info(store_lat, store_lon, last_shipment_lat, last_shipment_lon)
            # Ensure new trip still meets the distance constraints
            if new_trip_distance <= max_radius[vehicle_type]:
                new_offspring.append((vehicle_type, shipments, new_trip_distance, utilization, new_traffic))
                mutations_applied += 1
            else:
                new_offspring.append((vehicle_type, original_shipments, trip_distance, utilization, traffic))
                mutations_failed += 1
        else:
            new_offspring.append((vehicle_type, shipments, trip_distance, utilization, traffic))

    return new_offspring

max_vehicles_available = {"3W": 50, "4W-EV": 25, "4W": 10}
parent1 = selected_parents[0]
parent2 = selected_parents[1]
offspring = crossover_with_trip_selection(parent1, parent2, max_vehicles_available)
parent1_shp = sum(len(vehicle[1]) for vehicle in parent1)
print(f"Chromosome parent1 has {parent1_shp} shipments assigned.")
parent2_shp = sum(len(vehicle[1]) for vehicle in parent2)
print(f"Chromosome parent2 has {parent2_shp} shipments assigned.")
off_shp = sum(len(vehicle[1]) for vehicle in offspring)
print(f"Chromosome offspring has {off_shp} shipments assigned.")
print(offspring)
mutation_rate = 0.7
mutated_offspring = mutate_offspring(offspring, mutation_rate)
print(mutated_offspring)
total_shipments_assigned = sum(len(vehicle[1]) for vehicle in mutated_offspring)
print(f"Chromosome mutated offspring has {total_shipments_assigned} shipments assigned.")
print('Fitness of parent1: ', fitness_function(parent1))
print('Fitness of parent2: ', fitness_function(parent2))
print('Fitness of offspring: ', fitness_function(offspring))
print('Fitness of mutated offspring: ', fitness_function(mutated_offspring))
generations=10
for i in range(generations):
    print('Generation: ',i+1)
    selected_chromosomes=get_pareto_front(population)
    offspring_list = []
    for j in range(0, len(selected_chromosomes), 2):
        # Ensure the index does not go out of bounds for odd numbers of chromosomes
        if j + 1 < len(selected_chromosomes):
            parent1 = selected_chromosomes[j]
            parent2 = selected_chromosomes[j + 1]

            # Step 1: Apply crossover
            offspring = crossover_with_trip_selection(parent1, parent2, max_vehicles_available)

            # Step 2: Apply mutation to the offspring
            mutated_offspring = mutate_offspring(offspring, mutation_rate)

            # Step 3: Store the mutated offspring
            offspring_list.append(mutated_offspring)
    population=get_pareto_front(selected_chromosomes+offspring_list)
    for ind in population:
      print('Fitness: ',fitness_function(ind))
import numpy as np
avg_fitness=[]
for ind in population:
  avg_fitness.append(fitness_function(ind)[0]*0.6+fitness_function(ind)[1]*0.3+fitness_function(ind)[2]*0.1)
best_individual=population[np.argmax(avg_fitness)]
print(best_individual)
no_3w=0
no_4wev=0
no_4w=0
for gene in best_individual:
  if gene[0]=='3W':
    no_3w+=1
  elif gene[0]=='4W-EV':
    no_4wev+=1
  else:
    no_4w+=1
print(no_3w,no_4wev,no_4w)
print(best_individual)
from sklearn.cluster import KMeans
import numpy as np

def get_trip_coords(trip):
    """
    Extract the lat-long coordinates for a given trip.
    """
    return [(df.loc[df['Shipment ID'] == s, 'Latitude'].values[0],
             df.loc[df['Shipment ID'] == s, 'Longitude'].values[0]) for s in trip[1]]

def assign_to_existing_trip(trip, remaining_shipments, vehicle_type, max_capacity, max_radius):
    """
    Try to assign remaining shipments to an existing trip.
    """
    existing_shipments = trip[1]
    new_shipments = list(set(existing_shipments + remaining_shipments))  # Combine shipments

    if len(new_shipments) <= max_capacity[vehicle_type]:
        new_coords = get_trip_coords((vehicle_type, new_shipments, 0, 0))
        new_trip_distance = calculate_haversine_distance(new_coords)

        if new_trip_distance <= max_radius[vehicle_type]:
            return new_shipments, new_trip_distance
    return None, None

def create_new_trip(remaining_shipments, vehicle_type, max_capacity, max_radius):
    """
    Create a new trip for remaining shipments.
    """
    new_coords = get_trip_coords((vehicle_type, remaining_shipments, 0, 0))
    new_trip_distance = calculate_haversine_distance(new_coords)

    if len(remaining_shipments) <= max_capacity[vehicle_type] and new_trip_distance <= max_radius[vehicle_type]:
        return remaining_shipments, new_trip_distance
    return None, None

def minimize_4w_usage(offspring, max_vehicles):
    """
    Minimize the number of 4W vehicles by reassigning shipments and creating new trips from remaining vehicles.
    """
    max_capacity = {'3W': 5, '4W-EV': 8, '4W': 25}
    max_radius = {'3W': 15, '4W-EV': 20, '4W': float('inf')}

    remaining_shipments_3w = []
    remaining_shipments_4wev = []
    remaining_shipments_4w = []

    new_trips = []

    # Step 1: Try to reassign to 3W and 4W-EV vehicles first
    for trip in offspring:
        vehicle_type, shipments, trip_distance, utilization, traffic = trip

        if vehicle_type == '3W':
            remaining_shipments_3w.extend(shipments)
        elif vehicle_type == '4W-EV':
            remaining_shipments_4wev.extend(shipments)
        else:
            remaining_shipments_4w.extend(shipments)

    # Try to reassign shipments to 3W and 4W-EV vehicles
    for trip in offspring:
        vehicle_type, shipments, trip_distance, utilization, traffic = trip
        if vehicle_type == '3W':
            new_shipments, new_trip_distance = assign_to_existing_trip(trip, remaining_shipments_3w, '3W', max_capacity, max_radius)
            if new_shipments:
                new_traffic = get_traffic_info(store_lat, store_lon,
                                                df.loc[df['Shipment ID'] == new_shipments[-1], 'Latitude'].values[0],
                                                df.loc[df['Shipment ID'] == new_shipments[-1], 'Longitude'].values[0])
                new_trips.append(('3W', new_shipments, new_trip_distance, len(new_shipments)/max_capacity['3W'], new_traffic))
                remaining_shipments_3w = [s for s in remaining_shipments_3w if s not in new_shipments]

        elif vehicle_type == '4W-EV':
            new_shipments, new_trip_distance = assign_to_existing_trip(trip, remaining_shipments_4wev, '4W-EV', max_capacity, max_radius)
            if new_shipments:
                new_traffic = get_traffic_info(store_lat, store_lon,
                                                df.loc[df['Shipment ID'] == new_shipments[-1], 'Latitude'].values[0],
                                                df.loc[df['Shipment ID'] == new_shipments[-1], 'Longitude'].values[0])
                new_trips.append(('4W-EV', new_shipments, new_trip_distance, len(new_shipments)/max_capacity['4W-EV'], new_traffic))
                remaining_shipments_4wev = [s for s in remaining_shipments_4wev if s not in new_shipments]

    # Step 2: Create new trips from remaining shipments
    if remaining_shipments_3w:
        new_shipments, new_trip_distance = create_new_trip(remaining_shipments_3w, '3W', max_capacity, max_radius)
        if new_shipments:
            new_traffic = get_traffic_info(store_lat, store_lon,
                                            df.loc[df['Shipment ID'] == new_shipments[-1], 'Latitude'].values[0],
                                            df.loc[df['Shipment ID'] == new_shipments[-1], 'Longitude'].values[0])
            new_trips.append(('3W', new_shipments, new_trip_distance, len(new_shipments)/max_capacity['3W'], new_traffic))

    if remaining_shipments_4wev:
        new_shipments, new_trip_distance = create_new_trip(remaining_shipments_4wev, '4W-EV', max_capacity, max_radius)
        if new_shipments:
            new_traffic = get_traffic_info(store_lat, store_lon,
                                            df.loc[df['Shipment ID'] == new_shipments[-1], 'Latitude'].values[0],
                                            df.loc[df['Shipment ID'] == new_shipments[-1], 'Longitude'].values[0])
            new_trips.append(('4W-EV', new_shipments, new_trip_distance, len(new_shipments)/max_capacity['4W-EV'], new_traffic))

    # Step 3: Reassign any remaining shipments to 4W vehicles
    if remaining_shipments_4w:
        new_shipments, new_trip_distance = create_new_trip(remaining_shipments_4w, '4W', max_capacity, max_radius)
        if new_shipments:
            new_traffic = get_traffic_info(store_lat, store_lon,
                                            df.loc[df['Shipment ID'] == new_shipments[-1], 'Latitude'].values[0],
                                            df.loc[df['Shipment ID'] == new_shipments[-1], 'Longitude'].values[0])
            new_trips.append(('4W', new_shipments, new_trip_distance, len(new_shipments)/max_capacity['4W'], new_traffic))

    # Combine new trips with the original trips
    final_trips = new_trips + [trip for trip in offspring if trip[0] != '4W']

    return final_trips

best_individual=minimize_4w_usage(best_individual,max_vehicles_available)
print(fitness_function(best_individual))
no_3w=0
no_4wev=0
no_4w=0
for gene in best_individual:
  if gene[0]=='3W':
    no_3w+=1
  elif gene[0]=='4W-EV':
    no_4wev+=1
  else:
    no_4w+=1
print(best_individual)
print(no_3w,no_4wev,no_4w)
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

trip_counter = 101
all_trips_data = []

# Process each mutated_offspring trip
for offspring in mutated_offspring:
    vehicle_type = offspring[0]
    shipment_ids = offspring[1]
    mst_distance = offspring[2]
    capacity_uti = offspring[3]
    trip_time = round(mst_distance * 5)  # 5 min per km
    time_uti = 1 
    distance_uti = 1 
    trip_id = f"T{trip_counter}"  # Generate sequential trip ID

    # Fetch shipment details
    filtered_df = df[df["Shipment ID"].isin(shipment_ids)].copy()
    filtered_df["TRIP ID"] = trip_id
    filtered_df["Shipments"] = len(shipment_ids)
    filtered_df["MST_DIST"] = mst_distance
    filtered_df["TRIP_TIME"] = trip_time
    filtered_df["Vehical_Type"] = vehicle_type
    filtered_df["CAPACITY_UTI"] = capacity_uti
    filtered_df["TIME_UTI"] = time_uti
    filtered_df["COV_UTI"] = distance_uti

    # Append processed trip data
    all_trips_data.append(filtered_df)

    # Increment trip counter for next trip
    trip_counter += 1

# Combine all trips into a single DataFrame
final_df = pd.concat(all_trips_data, ignore_index=True)

# Rearrange columns
final_columns = ["TRIP ID", "Shipment ID", "Latitude", "Longitude", "Delivery Timeslot", "Shipments",
                 "MST_DIST", "TRIP_TIME", "Vehical_Type", "CAPACITY_UTI", "TIME_UTI", "COV_UTI"]
final_df = final_df[final_columns]

# Save to Excel
excel_filename = "optimized_trips.xlsx"
final_df.to_excel(excel_filename, index=False)

# Load Excel file for formatting
wb = load_workbook(excel_filename)
ws = wb.active

# Columns to Merge (Except "Shipment ID", "Latitude", "Longitude", "TIME SLOT")
merge_columns = ["A", "F", "G", "H", "I", "J", "K", "L"]  # TRIP ID, Shipments, MST_DIST, etc.
row_start = 2  # Start from row 2 (row 1 is header)
trip_ids = final_df["TRIP ID"].tolist()  # Use final_df instead of df

prev_trip = trip_ids[0]
merge_start = row_start

for i, trip_id in enumerate(trip_ids, start=row_start):
    if trip_id != prev_trip or i == len(trip_ids) + row_start - 1:  
        for col in merge_columns:
            ws.merge_cells(start_row=merge_start, start_column=ws[col + str(merge_start)].column, 
                           end_row=i - 1, end_column=ws[col + str(merge_start)].column)
            ws[f"{col}{merge_start}"].alignment = Alignment(horizontal="center", vertical="center")
        merge_start = i  # Update start position
    prev_trip = trip_id

# Save the formatted Excel file
wb.save(excel_filename)
print(f"Formatted Excel file saved as: {excel_filename}")
